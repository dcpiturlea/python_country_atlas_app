import openpyxl

def open_xl(wb_path):
    wb = openpyxl.load_workbook(wb_path)
    main_sheet = wb.active
    return wb


def countries_list(wb):
    worksheet_names = wb.sheetnames
    sheet_index = worksheet_names.index("Sheet1")
    wb.active = sheet_index
    sheet_keyword = wb.active

    last_column = sheet_keyword.max_column
    last_row = sheet_keyword.max_row
    countries = {}

    for i in range(1, last_row):
        country = sheet_keyword.cell(row=i, column=1).value
        countries[i] = country
    return countries

